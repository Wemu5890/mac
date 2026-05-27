import os
import sys
import uuid
import tempfile
import threading
import webbrowser
import socket
import json
import urllib.request
import urllib.error
import ssl  
import subprocess
import openpyxl
import pandas as pd
import re
from flask import Flask, request, jsonify, send_file, send_from_directory

def get_base_path():
    """获取程序运行时的绝对路径（兼容 PyInstaller 单文件解压路径）"""
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.abspath(".")

base_path = get_base_path()

app = Flask(__name__, static_folder=base_path, static_url_path='')
app.config['UPLOAD_FOLDER'] = tempfile.gettempdir()

SESSION_FILES = {}

HEADER_ALIASES = {
    "所教学科": ["学科", "科目", "任教学科", "任教科目", "教学科目", "教学科"],
    "编号": ["序号", "工号", "教师编号", "ID", "教师序号", "教职工号"],
    "姓名": ["老师", "教师姓名", "教师", "名字"],
    "联系方式": ["电话", "手机", "手机号", "联系电话", "联系号码", "手机号码"],
    "身份证号": ["身份证", "证件号", "身份证号码"]
}

def standardize_header(raw_header):
    if not raw_header:
        return ""
    raw = str(raw_header).strip()
    for std, aliases in HEADER_ALIASES.items():
        if raw == std: return std
    for std, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in raw:
                return std
    return raw

ssl_context = ssl._create_unverified_context()

@app.route('/api/check_update')
def check_update():
    try:
        req = urllib.request.Request("https://api.github.com/repos/Wemu5890/mac/releases/latest", headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10, context=ssl_context) as response:
            data = json.loads(response.read().decode('utf-8'))
            latest_version = data.get('tag_name', '')
            download_url = ''
            if 'assets' in data and len(data['assets']) > 0:
                is_mac = sys.platform == 'darwin'
                for asset in data['assets']:
                    name = asset['name'].lower()
                    if is_mac and ('mac' in name or name.endswith('.zip') or name.endswith('.dmg')):
                        download_url = asset['browser_download_url']
                        break
                    elif not is_mac and name.endswith('.exe'):
                        download_url = asset['browser_download_url']
                        break
            
            has_update = False
            if latest_version:
                has_update = latest_version.replace('v', '') > "1.0.31"
            
            return jsonify({
                "current_version": "v1.0.31",
                "latest_version": latest_version,
                "has_update": has_update,
                "download_url": download_url,
                "release_notes": data.get('body', '')
            })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/do_update', methods=['POST'])
def do_update():
    data = request.json
    download_url = data.get('download_url')
    if not download_url:
        return jsonify({"error": "缺少下载链接"}), 400
    try:
        is_mac = sys.platform == 'darwin'
        ext = ".zip" if is_mac else ".exe"
        installer_path = os.path.join(tempfile.gettempdir(), f"update_package{ext}")
        
        req = urllib.request.Request(download_url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=60, context=ssl_context) as response, open(installer_path, 'wb') as out_file:
            out_file.write(response.read())
            
        if is_mac:
            subprocess.Popen(['open', installer_path])
            threading.Timer(0.5, lambda: os._exit(0)).start()
        elif os.name == 'nt':
            subprocess.Popen([installer_path, '/SILENT', '/SP-'], creationflags=subprocess.CREATE_NEW_CONSOLE | subprocess.CREATE_NEW_PROCESS_GROUP)
            threading.Timer(0.5, lambda: os._exit(0)).start()
        return jsonify({"message": "更新下载完成，即将自动重启覆盖安装！"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/')
def index():
    return send_from_directory(base_path, 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory(base_path, path)

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "没有收到文件"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "未选择文件"}), 400
        
    session_id = str(uuid.uuid4())
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{file.filename}")
    file.save(file_path)
    SESSION_FILES[session_id] = file_path
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        headerRowIndex = -1
        headerRowData = []
        bodyData = []
        
        for row in ws.iter_rows():
            found = False
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                if '编号' in val or '姓名' in val or '工号' in val:
                    found = True
                    break
            if found:
                headerRowIndex = row[0].row
                headerRowData = [""] * (ws.max_column + 1)
                for cell in row:
                    headerRowData[cell.column] = standardize_header(cell.value) if cell.value is not None else ""
                break
                
        if headerRowIndex == -1:
            return jsonify({"error": "未能识别出包含“编号”或“姓名”的表头行，请检查文件格式"}), 400
            
        idIndex = -1
        nameIndex = -1
        for i, h in enumerate(headerRowData):
            if '编号' in h: idIndex = i
            if '姓名' in h: nameIndex = i
            
        for row in ws.iter_rows(min_row=headerRowIndex + 1):
            rowArr = [""] * (ws.max_column + 1)
            for cell in row:
                if cell.value is not None:
                    if hasattr(cell.value, 'strftime'):
                        rowArr[cell.column] = cell.value.strftime('%Y/%m/%d')
                    else:
                        rowArr[cell.column] = str(cell.value).strip()
            
            valId = rowArr[idIndex] if idIndex >= 0 else ""
            valName = rowArr[nameIndex] if nameIndex >= 0 else ""
            hasId = bool(valId and valId.strip())
            hasName = bool(valName and valName.strip())
            
            if idIndex == -1 and nameIndex == -1:
                if any(c.strip() for c in rowArr):
                    bodyData.append({"excelRowNumber": row[0].row, "cells": rowArr})
            elif hasId or hasName:
                bodyData.append({"excelRowNumber": row[0].row, "cells": rowArr})
                
        return jsonify({
            "sessionId": session_id,
            "headerRowData": headerRowData,
            "bodyData": bodyData
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/export', methods=['POST'])
def export_file():
    data = request.json
    session_id = data.get('sessionId')
    updates = data.get('updates', [])
    
    if not session_id or session_id not in SESSION_FILES:
        return jsonify({"error": "会话无效或文件已过期，请重新上传！"}), 400
        
    file_path = SESSION_FILES[session_id]
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 【核心紧跟策略】：动态扫描 Excel 中真正有文字数据的最后一行的位置
        real_max_row = 1
        for row_idx in range(1, ws.max_row + 1):
            if any(ws.cell(row=row_idx, column=col_idx).value is not None for col_idx in range(1, ws.max_column + 1)):
                real_max_row = row_idx

        row_updates = {}
        for update in updates:
            raw_row = update.get('excelRowNumber')
            try:
                c = int(update.get('colNum', 0))
            except Exception:
                continue
            if c < 1:
                continue
                
            val = update.get('value', '')
            if val.strip() and val.strip().replace('.', '', 1).isdigit():
                try:
                    val = float(val) if '.' in val else int(val)
                except ValueError:
                    pass
            
            try:
                r = int(float(raw_row)) if raw_row is not None and str(raw_row).strip() not in ['', 'null', 'undefined', 'None'] else -1
            except Exception:
                r = -1
                
            if r == 0:
                continue
                
            if r not in row_updates:
                row_updates[r] = []
            row_updates[r].append((c, val))
            
        # 按照负数虚拟行号排序，确保新追加的人不会乱
        sorted_rows = sorted(row_updates.keys(), key=lambda x: (0, x) if x >= 0 else (1, -x))
        
        current_append_row = real_max_row + 1
        for r in sorted_rows:
            cells = row_updates[r]
            if r < 0:
                # 抛弃 ws.append 隐形空行大雷，采用高精度底层写值，写完行号顺序下移一行
                for c, val in cells:
                    ws.cell(row=current_append_row, column=c, value=val)
                current_append_row += 1
            else:
                for c, val in cells:
                    ws.cell(row=r, column=c, value=val)
            
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], f"updated_{session_id}.xlsx")
        wb.save(output_path)
        
        return send_file(
            output_path, 
            as_attachment=True, 
            download_name="已更新数据.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/api/process_students', methods=['POST'])
def process_students():
    if 'template' not in request.files or 'master' not in request.files or 'classlist' not in request.files:
        return jsonify({"error": "缺失必要文件"}), 400
        
    tmpl_file = request.files['template']
    master_file = request.files['master']
    class_file = request.files['classlist']
    
    session_id = str(uuid.uuid4())
    tmpl_path = os.path.join(app.config['UPLOAD_FOLDER'], f"stu_{session_id}_tmpl.xlsx")
    master_path = os.path.join(app.config['UPLOAD_FOLDER'], f"stu_{session_id}_master.tmp")
    class_path = os.path.join(app.config['UPLOAD_FOLDER'], f"stu_{session_id}_class.tmp")
    
    tmpl_file.save(tmpl_path)
    master_file.save(master_path)
    class_file.save(class_path)
    
    try:
        def get_clean_df(filepath):
            """智能读取表格，自动跳过头上多余的非表头空行"""
            df = pd.read_excel(filepath, header=None)
            header_idx = -1
            for i, row in df.iterrows():
                if any(isinstance(val, str) and '姓名' in val for val in row.values):
                    header_idx = i
                    break
            if header_idx != -1:
                df.columns = [str(c).strip() for c in df.iloc[header_idx]]
                df = df.iloc[header_idx+1:]
            return df
            
        # 1. 解析学生总表，建立【姓名 -> 4位考号】的高速索引字典
        df_master = get_clean_df(master_path)
        name_col_m = next((c for c in df_master.columns if '姓名' in c), None)
        class_col_m = next((c for c in df_master.columns if '行政班' in c or '班级' in c or c == '班'), None)
        id_col_m = next((c for c in df_master.columns if '学号' in c), None)
        
        student_map = {}
        if name_col_m and class_col_m and id_col_m:
            for _, row in df_master.iterrows():
                name = str(row[name_col_m]).strip()
                cls_val = str(row[class_col_m]).strip()
                id_val = str(row[id_col_m]).strip()
                
                # 纯净提取数字
                cls_nums = re.findall(r'\d+', cls_val)
                id_nums = re.findall(r'\d+', id_val)
                
                if name and name != 'nan' and cls_nums and id_nums:
                    # 核心规则：班级在前，学号在后，强制补零为2位
                    exam_num = f"{int(cls_nums[0]):02d}{int(id_nums[0]):02d}"
                    student_map[name] = exam_num
                    
        # 2. 解析老师提供的名单表 (支持不规范的 .xls 格式)
        df_class = get_clean_df(class_path)
        name_col_c = next((c for c in df_class.columns if '姓名' in c), None)
        class_col_c = next((c for c in df_class.columns if '教学班' in c or '分班' in c or '目标班' in c), None)
        
        target_students = []
        if name_col_c:
            for _, row in df_class.iterrows():
                name = str(row[name_col_c]).strip()
                if not name or name == 'nan': continue
                tgt_class = str(row[class_col_c]).strip() if class_col_c else ""
                target_students.append((name, tgt_class))
                
        # 3. 填入导入模板并导出
        wb = openpyxl.load_workbook(tmpl_path)
        ws = wb.active
        
        t_name_idx = -1
        t_class_idx = -1
        t_exam_idx = -1
        
        header_row_idx = 1
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and '姓名' in str(cell.value):
                    header_row_idx = row[0].row
                    break
            if header_row_idx > 1 or (row[0].value and '姓名' in str(row[0].value)): break
            
        for cell in ws[header_row_idx]:
            val = str(cell.value).strip() if cell.value else ""
            if '姓名' in val: t_name_idx = cell.column
            elif '教学班' in val or '班级' in val: t_class_idx = cell.column
            elif '考号' in val or '考试号' in val: t_exam_idx = cell.column
            
        current_row = header_row_idx + 1
        for name, tgt_class in target_students:
            exam_num = student_map.get(name, "")
            if t_name_idx > 0: ws.cell(row=current_row, column=t_name_idx, value=name)
            if t_class_idx > 0 and tgt_class and tgt_class != 'nan': 
                ws.cell(row=current_row, column=t_class_idx, value=tgt_class)
            if t_exam_idx > 0: ws.cell(row=current_row, column=t_exam_idx, value=exam_num)
            current_row += 1
            
        # 填入完毕后，将其直接存为本次的 Session 模板文件，以便前端使用已有的导出/复查通道
        wb.save(tmpl_path)
        SESSION_FILES[session_id] = tmpl_path
        
        # 提取数据供前端复查（和 /api/upload 保持一致的数据结构）
        headerRowData = [""] * (ws.max_column + 1)
        for cell in ws[header_row_idx]:
            headerRowData[cell.column] = standardize_header(cell.value) if cell.value is not None else ""
            
        bodyData = []
        for row in ws.iter_rows(min_row=header_row_idx + 1):
            rowArr = [""] * (ws.max_column + 1)
            for cell in row:
                if cell.value is not None:
                    if hasattr(cell.value, 'strftime'):
                        rowArr[cell.column] = cell.value.strftime('%Y/%m/%d')
                    else:
                        rowArr[cell.column] = str(cell.value).strip()
            
            if any(c.strip() for c in rowArr):
                bodyData.append({"excelRowNumber": row[0].row, "cells": rowArr})
                
        return jsonify({
            "sessionId": session_id,
            "headerRowIndex": header_row_idx,
            "headerRowData": headerRowData,
            "bodyData": bodyData
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

def find_free_port():
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(('', 0))
        s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        return s.getsockname()[1]

def open_as_app(url):
    try:
        if sys.platform == 'win32':
            cmd = f'start "" chrome --app="{url}" || start "" msedge --app="{url}"'
            subprocess.run(cmd, shell=True)
        elif sys.platform == 'darwin':
            cmd = f'open -n -a "Google Chrome" --args --app="{url}" || open -n -a "Microsoft Edge" --args --app="{url}"'
            subprocess.run(cmd, shell=True)
        else:
            webbrowser.open_new(url)
    except Exception:
        webbrowser.open_new(url)

if __name__ == '__main__':
    port = find_free_port()
    url = f"http://127.0.0.1:{port}"
    threading.Timer(1.5, lambda: open_as_app(url)).start()
    app.run(host='127.0.0.1', port=port, debug=False)